# -*- coding: utf-8 -*-
import os
import sys
import argparse
import openpyxl
import xlrd
import xml.etree.ElementTree as ET

# 检查路径是否存在
def verify_path(path):
    if os.path.exists(path):
        return True
    raise Exception(f"❌ 文件或目录未找到: {path}")

# 获取项目根目录下，各模块英文string.xml文件绝对路径
def get_en_file_path(project_path, exclude_modules=None):
    if exclude_modules is None:
        exclude_modules = []
    
    # 各模块下的string_en.xml绝对路径
    en_xml_path_list = []
    # 各模块下res的绝对路径
    res_path_list = []
    
    print(f"正在扫描: {project_path}")
    if exclude_modules:
        print(f"忽略模块: {exclude_modules}")

    # 遍历目录树
    for root, dirs, files in os.walk(project_path):
        # 排除指定的模块（目录）
        # 修改 dirs 列表会影响 os.walk 的后续遍历，从而跳过这些目录
        for exclude in exclude_modules:
            if exclude in dirs:
                dirs.remove(exclude)
                print(f'\r已跳过忽略模块: {exclude}{" " * 30}')

        print(f'\r正在搜索: {root}{" " * 50}', end='', flush=True)
        
        # 检查当前目录是否是 res 目录
        if os.path.basename(root) == 'res':
            # 构建 values/strings.xml 路径
            strings_xml_path = os.path.join(root, 'values', 'strings.xml')
            # 构建 values/string.xml 路径
            string_xml_path = os.path.join(root, 'values', 'string.xml')
            
            # 检查文件是否存在
            found_path = None
            if os.path.isfile(string_xml_path):
                found_path = string_xml_path
            elif os.path.isfile(strings_xml_path):
                found_path = strings_xml_path
                
            if found_path:
                en_xml_path_list.append(found_path)
                res_path_list.append(root)
                print(f'\r找到英文文件，路径为：{found_path}{" " * 20}')
                
    print()
    return en_xml_path_list, res_path_list


# 解析英文的strings.xml
def parse_en_xml(en_xml_path: str, result_xml_path: str, translate_dict: dict, no_translate_list: list):
    try:
        tree = ET.parse(en_xml_path)
        root = tree.getroot()

        string_elements = root.findall('.//string')
        total_strings = len(string_elements)

        print(f"{en_xml_path} 共找到 {total_strings} 个string标签需要翻译")
        print("=" * 60)

        for i, string_elem in enumerate(string_elements, 1):
            if string_elem.text:
                string_id = string_elem.attrib['name']
                en_text = string_elem.text
                
                # 尝试翻译
                translate_text = translate_dict.get(string_id, "")
                
                print(f"正在翻译第 {i}/{total_strings} 个标签【{string_id}】")
                # print(f'原文：{en_text}')
                # print(f'译文: {translate_text}') # 减少刷屏，如需查看细节可取消注释
                # print("-" * 50)
                
                if translate_text == "":
                    no_translate_list.append([string_id, string_elem.text])
                    # 保留原文或者置空，视需求而定，这里保持原逻辑：赋空值给text可能会导致空标签
                    # 原逻辑是: string_elem.text = translate_text
                    # 如果翻译为空，标签内容变为空
                string_elem.text = translate_text

        # 写入结果
        tree.write(result_xml_path, encoding='utf-8', xml_declaration=True)
        print("=" * 60)
        print(f"翻译完成！翻译结果已保存到: {result_xml_path}")
        return True

    except ET.ParseError as e:
        print(f"XML解析错误: {e}")
        return False
    except Exception as e:
        print(f"处理文件时出错: {e}")
        return False


# 没有翻译的id，保存到xlsx文件中，表结构为：ID  EN
def no_translate_save_xml(no_translate_list):
    lack_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'lack_translate_id.xlsx')
    if os.path.exists(lack_file):
        try:
            os.remove(lack_file)
        except Exception:
            print(f"⚠️ 无法删除旧的 {lack_file}，请检查是否被占用")

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ['ID', 'EN']
    ws.append(headers)
    for row_data in no_translate_list:
        ws.append(row_data)
    wb.save(lack_file)
    print(f"⚠️ 无翻译数据ID已写入到 {lack_file} 文件中")


# 读取excel翻译文件
def read_translate_excel(translate_file_path, code=''):
    print(f"正在读取翻译文件: {translate_file_path} ...")
    translate_dict = {}
    
    # 尝试使用 xlrd 打开 (针对 .xls)
    # 注意：xlrd 新版不支持 .xlsx，如果遇到 .xlsx 需要用 openpyxl 读取
    # 这里为了兼容原代码库逻辑，保留 xlrd，但建议用户尽量转为 openpyxl
    try:
        wb = xlrd.open_workbook(translate_file_path)
        sheet = wb.sheet_by_index(0) # 默认读取第一个sheet
        
        # 寻找对应的列
        mycol = -1
        #表头
        header_row = sheet.row_values(0)
        for col_index, value in enumerate(header_row):
            if str(value).lower() == code.lower():
                mycol = col_index
                break
        
        if mycol == -1:
            raise Exception(f'⚠️ Excel中没有找到名为 {code} 的语言列')

        # 读取数据
        rows = sheet.nrows
        for row in range(1, rows): # 从第1行开始（跳过表头）
             key = sheet.cell(row, 0).value
             value = sheet.cell(row, mycol).value
             if key:
                 translate_dict[str(key)] = str(value)
                 
    except Exception as e:
        # 如果 xlrd 失败，且是 xlsx，提示错误
        print(f"读取Excel失败: {e}")
        raise e

    print(f"✅ 读取完成，共加载 {len(translate_dict)} 条翻译")
    return translate_dict


def main():
    parser = argparse.ArgumentParser(description="Android 项目国际化自动翻译脚本")
    parser.add_argument("--excel", help="翻译 Excel 文件路径 (.xls)")
    parser.add_argument("--project", help="Android 项目根目录路径")
    parser.add_argument("--lang", default="no", help="目标语言代码 (例如: no, fr, es)，默认: no")
    parser.add_argument("--exclude", nargs='+', help="需要排除的模块名 (目录名)，例如: build test values-xx", default=[])
    
    args = parser.parse_args()
    
    print("="*60)
    print("        Android 自动翻译脚本 v2.0 (Optimized)")
    print("="*60)

    # 如果没有通过命令行传入参数，则交互式输入
    if not args.excel or not args.project:
        print("💡 提示: 你也可以通过命令行参数运行此脚本，避免手动输入。")
        print("   示例: python android_translate_string_by_excel.py --excel 'E:/tr.xls' --project 'D:/App' --lang 'fr' --exclude 'lib_a' 'lib_b'")
        print("-" * 60)
        
        excel_path = args.excel if args.excel else input("请输入 Excel 翻译文件路径: ").strip().replace('"', '')
        project_path = args.project if args.project else input("请输入 Android 项目路径: ").strip().replace('"', '')

        if not args.excel: # 只有在没传参时才问语言
            lang_input = input(f"请输入目标语言代码 (默认 {args.lang}): ").strip()
            language_code = lang_input if lang_input else args.lang
        else:
            language_code = args.lang

        if not args.exclude:
            exclude_input = input("请输入要排除的模块名 (空格分隔，直接回车跳过): ").strip()
            exclude_modules = exclude_input.split() if exclude_input else []
        else:
            exclude_modules = args.exclude

    else:
        excel_path = args.excel
        project_path = args.project
        language_code = args.lang
        exclude_modules = args.exclude

    # 路径校验
    verify_path(excel_path)
    verify_path(project_path)
    
    values_folder_name = f'values-{language_code}'
    
    # 1. 读取 Excel (只读取一次!)
    translate_dict = read_translate_excel(excel_path, language_code)
    
    # 2. 获取所有目标 strings.xml
    en_xml_path_lists, res_path_lists = get_en_file_path(project_path, exclude_modules)
    
    if not en_xml_path_lists:
        print("⚠️ 未找到任何 strings.xml 文件，请检查项目路径或排除规则。")
        return

    no_translate_id_list = [] # 用于收集无翻译的条目
    
    # 3. 遍历处理每个文件
    for i in range(len(en_xml_path_lists)):
        final_translate_dir = os.path.join(res_path_lists[i], values_folder_name)
        
        if not os.path.exists(final_translate_dir):
            os.mkdir(final_translate_dir)
            
        final_translate_path = os.path.join(final_translate_dir, 'strings.xml')
        
        # 解析并替换
        parse_en_xml(en_xml_path_lists[i], final_translate_path, translate_dict, no_translate_id_list)

    # 4. 保存未翻译条目
    if no_translate_id_list:
        no_translate_save_xml(no_translate_id_list)
    else:
        print("🎉 完美！所有条目都已找到对应翻译。")
        
    print('✅ 程序结束！')


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding='utf-8')
    try:
        main()
    except KeyboardInterrupt:
        print("\n❌ 用户中断执行")
    except Exception as e:
        print(f"\n❌ 程序发生错误: {e}")
        import traceback
        traceback.print_exc()
